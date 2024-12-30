import openpyxl
import os

def analyze_excel(file_path, output_file):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path, data_only=True)

    with open(output_file, 'w') as f:
        # Write the name of the Excel file
        f.write(f"Excel File Name: {os.path.basename(file_path)}\n")

        # Write the number of sheets
        f.write(f"Number of Sheets: {len(wb.sheetnames)}\n")

        # Write the names of the sheets
        f.write("Sheet Names:\n")
        for sheet_name in wb.sheetnames:
            f.write(f"- {sheet_name}\n")

        # Analyze each sheet
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            f.write(f"\nAnalyzing Sheet: {sheet_name}\n")

            # Get column names (assuming first row contains headers)
            column_names = [cell.value for cell in sheet[1]]
            f.write("Column Names:\n")
            for col_name in column_names:
                f.write(f"- {col_name}\n")

            # Write sample of first two rows after column names
            f.write("\nSample Data (First Two Rows):\n")
            for row in sheet.iter_rows(min_row=2, max_row=3, values_only=True):
                f.write(f"{row}\n")

    wb.close()
    print(f"Analysis complete. Results saved to {output_file}")

# Example usage
file_path = "studio_database_1.xlsx"
output_file = "excel_notes_2.txt"
analyze_excel(file_path, output_file)
